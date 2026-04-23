#!/usr/bin/env python3
"""
Comprehensive Experiment Analyzer for FedDSA Project
Reads ALL JSON result files, extracts metrics, maps to EXP numbers, outputs Excel.
"""

import json
import os
import re
import glob
import numpy as np
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TASK_DIRS = {
    "PACS_c4": os.path.join(BASE_DIR, "FDSE_CVPR25", "task", "PACS_c4", "record"),
    "office_caltech10_c4": os.path.join(BASE_DIR, "FDSE_CVPR25", "task", "office_caltech10_c4", "record"),
    "domainnet_c6": os.path.join(BASE_DIR, "FDSE_CVPR25", "task", "domainnet_c6", "record"),
}

# Client names per dataset
CLIENT_NAMES = {
    "PACS_c4": ["Art_painting", "Cartoon", "Photo", "Sketch"],
    "office_caltech10_c4": ["Amazon", "Caltech", "DSLR", "Webcam"],
    "domainnet_c6": ["Clipart", "Infograph", "Painting", "Quickdraw", "Real", "Sketch"],
}

EXP_DIR = os.path.join(BASE_DIR, "experiments")


def extract_seed_from_filename(filename):
    """Extract seed from flgo filename like ..._S2_LD0_..."""
    m = re.search(r'_S(\d+)_LD', filename)
    if m:
        return int(m.group(1))
    return None


def extract_seed_from_json(data):
    """Extract seed from JSON option dict."""
    opt = data.get("option", {})
    return opt.get("seed", None)


def extract_algo_from_filename(filename):
    """Extract algorithm name from filename."""
    basename = os.path.basename(filename)
    # Remove .json extension
    name = basename.replace(".json", "")

    # Known algorithm prefixes (order matters - longer first)
    algo_prefixes = [
        "feddsa_adaptive", "feddsa_augschedule", "feddsa_auto",
        "feddsa_cka", "feddsa_consensus", "feddsa_domain_aware",
        "feddsa_fixbn", "feddsa_gated", "feddsa_gradual",
        "feddsa_multilayer", "feddsa_noaug_late", "feddsa_noaug",
        "feddsa_pcgrad", "feddsa_regime_gated_v2", "feddsa_regime_gated",
        "feddsa_softbeta", "feddsa_stylehead_bank", "feddsa_triplet",
        "feddsa_vae", "feddsa_plus", "feddsa_stable",
        "feddsa",
        "fdse", "fedavg", "fedbn", "fedproto", "fedprox",
        "moon", "ditto",
    ]

    for prefix in algo_prefixes:
        if name.startswith(prefix + "_"):
            return prefix
    return name.split("_")[0]


def extract_params_signature(filename):
    """Extract a parameter signature string from filename for grouping."""
    basename = os.path.basename(filename).replace(".json", "")
    # Remove seed part and after
    # Pattern: ..._S{seed}_LD0_...
    m = re.search(r'_S\d+_LD', basename)
    if m:
        return basename[:m.start()]
    return basename


def parse_json_file(filepath):
    """Parse a single JSON result file and extract all metrics."""
    try:
        with open(filepath, "r") as f:
            data = json.load(f)
    except (json.JSONDecodeError, FileNotFoundError) as e:
        print(f"  ERROR reading {filepath}: {e}")
        return None

    result = {}
    basename = os.path.basename(filepath)
    result["filename"] = basename
    result["filepath"] = filepath

    # Extract option/config
    opt = data.get("option", {})
    result["seed"] = opt.get("seed", extract_seed_from_filename(basename))
    result["learning_rate"] = opt.get("learning_rate", None)
    result["num_rounds"] = opt.get("num_rounds", None)
    result["num_epochs"] = opt.get("num_epochs", None)
    result["batch_size"] = opt.get("batch_size", None)
    result["weight_decay"] = opt.get("weight_decay", None)
    result["lr_decay"] = opt.get("learning_rate_decay", None)
    result["algo_para"] = opt.get("algo_para", None)
    result["algorithm"] = extract_algo_from_filename(basename)

    # Extract parameter signature for grouping same-config runs
    result["param_sig"] = extract_params_signature(basename)

    # ---- Accuracy metrics ----
    # local_val_accuracy: weighted mean val accuracy per round
    local_val_acc = data.get("local_val_accuracy", [])
    local_test_acc = data.get("local_test_accuracy", [])
    mean_local_test_acc = data.get("mean_local_test_accuracy", [])
    std_local_test_acc = data.get("std_local_test_accuracy", [])

    result["total_rounds"] = len(local_val_acc)

    if local_val_acc:
        arr = np.array(local_val_acc)
        result["max_val_acc"] = float(arr.max())
        result["best_val_round"] = int(arr.argmax())
        result["last_val_acc"] = float(arr[-1])
        result["val_gap"] = float(arr.max() - arr[-1])
    else:
        result["max_val_acc"] = None
        result["best_val_round"] = None
        result["last_val_acc"] = None
        result["val_gap"] = None

    if local_test_acc:
        arr = np.array(local_test_acc)
        result["max_test_acc"] = float(arr.max())
        result["best_test_round"] = int(arr.argmax())
        result["last_test_acc"] = float(arr[-1])
        result["test_gap"] = float(arr.max() - arr[-1])

        # Test acc at best val round
        if result["best_val_round"] is not None and result["best_val_round"] < len(local_test_acc):
            result["test_at_best_val"] = float(local_test_acc[result["best_val_round"]])
        else:
            result["test_at_best_val"] = None
    else:
        result["max_test_acc"] = None
        result["best_test_round"] = None
        result["last_test_acc"] = None
        result["test_gap"] = None
        result["test_at_best_val"] = None

    if mean_local_test_acc:
        arr = np.array(mean_local_test_acc)
        result["max_mean_test_acc"] = float(arr.max())
        # Mean test at best val round
        if result["best_val_round"] is not None and result["best_val_round"] < len(mean_local_test_acc):
            result["mean_test_at_best_val"] = float(mean_local_test_acc[result["best_val_round"]])
        else:
            result["mean_test_at_best_val"] = None
    else:
        result["max_mean_test_acc"] = None
        result["mean_test_at_best_val"] = None

    # ---- Per-client accuracy at best val round ----
    test_dist = data.get("local_test_accuracy_dist", [])
    if test_dist and result["best_val_round"] is not None:
        best_round = result["best_val_round"]
        if best_round < len(test_dist):
            client_accs = test_dist[best_round]
            result["num_clients"] = len(client_accs)
            result["per_client_acc_at_best"] = [float(x) for x in client_accs]
        else:
            result["num_clients"] = len(test_dist[0]) if test_dist else 0
            result["per_client_acc_at_best"] = []
    else:
        result["num_clients"] = 0
        result["per_client_acc_at_best"] = []

    # Per-client best (each client's best across all rounds, selected by that client's val)
    val_dist = data.get("local_val_accuracy_dist", [])
    if test_dist and val_dist:
        n_clients = len(test_dist[0]) if test_dist else 0
        per_client_best = []
        for i in range(n_clients):
            client_val = [val_dist[r][i] for r in range(len(val_dist))]
            client_test = [test_dist[r][i] for r in range(len(test_dist))]
            best_r = int(np.argmax(client_val))
            per_client_best.append(float(client_test[best_r]))
        result["per_client_best_acc"] = per_client_best
    else:
        result["per_client_best_acc"] = []

    # ---- Per-client accuracy at last round ----
    if test_dist:
        result["per_client_acc_last"] = [float(x) for x in test_dist[-1]]
    else:
        result["per_client_acc_last"] = []

    return result


def read_note_md(filepath):
    """Read a NOTE.md file and extract key info."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
        # Extract first few lines as summary
        lines = content.strip().split("\n")
        title = ""
        description = ""
        for line in lines[:30]:
            if line.startswith("# "):
                title = line[2:].strip()
            elif line.startswith("## ") and not description:
                pass
            elif line.strip() and not title:
                title = line.strip()
        return {"title": title, "content": content[:2000], "full": content}
    except Exception:
        return {"title": "", "content": "", "full": ""}


def scan_all_experiments():
    """Scan experiments/ directory for all EXP folders and their NOTE.md."""
    exp_info = {}
    for note_path in glob.glob(os.path.join(EXP_DIR, "**", "NOTE.md"), recursive=True):
        # Extract EXP number
        dirname = os.path.basename(os.path.dirname(note_path))
        m = re.match(r'EXP-(\d+)', dirname)
        if m:
            exp_num = int(m.group(1))
            note_data = read_note_md(note_path)
            exp_info[exp_num] = {
                "exp_id": f"EXP-{exp_num:03d}",
                "folder_name": dirname,
                "folder_path": os.path.dirname(note_path),
                "note_title": note_data["title"],
                "note_content": note_data["content"],
                "note_full": note_data["full"],
            }
    return exp_info


def scan_all_json_results():
    """Scan all task record directories for JSON result files."""
    all_results = {}
    for task_name, task_dir in TASK_DIRS.items():
        if not os.path.exists(task_dir):
            print(f"  WARNING: {task_dir} does not exist")
            continue
        json_files = sorted(glob.glob(os.path.join(task_dir, "*.json")))
        task_results = []
        for jf in json_files:
            print(f"  Parsing {task_name}: {os.path.basename(jf)[:60]}...")
            r = parse_json_file(jf)
            if r:
                r["task"] = task_name
                task_results.append(r)
        all_results[task_name] = task_results
        print(f"  {task_name}: {len(task_results)} results parsed")
    return all_results


def group_by_config(results_list):
    """Group results by parameter signature (same config, different seeds)."""
    groups = defaultdict(list)
    for r in results_list:
        groups[r["param_sig"]].append(r)
    return dict(groups)


def compute_multi_seed_stats(group):
    """Compute mean/std across seeds for a group of runs."""
    if not group:
        return {}
    stats = {}
    metrics = ["max_val_acc", "max_test_acc", "max_mean_test_acc",
               "test_at_best_val", "mean_test_at_best_val",
               "last_val_acc", "last_test_acc", "val_gap", "test_gap",
               "best_val_round"]
    for m in metrics:
        vals = [r[m] for r in group if r[m] is not None]
        if vals:
            stats[f"{m}_mean"] = float(np.mean(vals))
            stats[f"{m}_std"] = float(np.std(vals))
            stats[f"{m}_min"] = float(np.min(vals))
            stats[f"{m}_max"] = float(np.max(vals))
        else:
            stats[f"{m}_mean"] = None
            stats[f"{m}_std"] = None
    stats["n_seeds"] = len(group)
    stats["seeds"] = sorted([r["seed"] for r in group if r["seed"] is not None])
    return stats


# ============= Excel Output =============

def pct(val):
    """Format as percentage string."""
    if val is None:
        return ""
    return round(val * 100, 2)


def pct_str(val):
    if val is None:
        return ""
    return f"{val*100:.2f}%"


def mean_std_str(mean, std):
    if mean is None:
        return ""
    if std is not None and std > 0:
        return f"{mean*100:.2f}±{std*100:.2f}"
    return f"{mean*100:.2f}"


def write_excel(all_results, exp_info, output_path):
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row=1, max_col=None):
        if max_col is None:
            max_col = ws.max_column
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def auto_width(ws, min_width=8, max_width=40):
        for col_cells in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
            ws.column_dimensions[col_letter].width = max_len + 2

    # ==========================================
    # Sheet 1: Overview - All Individual Runs
    # ==========================================
    ws1 = wb.active
    ws1.title = "All Runs"

    headers = [
        "Task", "Algorithm", "Seed", "LR", "Rounds", "Epochs",
        "Max Val Acc%", "Best Val Round", "Last Val Acc%", "Val Gap%",
        "Max Test Acc%", "Best Test Round", "Last Test Acc%", "Test Gap%",
        "Test@BestVal%", "Mean Test@BestVal%", "Max Mean Test%",
        "Param Signature"
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, max_col=len(headers))

    row = 2
    for task_name in ["PACS_c4", "office_caltech10_c4", "domainnet_c6"]:
        results = all_results.get(task_name, [])
        # Sort by algorithm then seed
        results.sort(key=lambda x: (x["algorithm"], x.get("seed", 0) or 0))
        for r in results:
            ws1.cell(row=row, column=1, value=task_name)
            ws1.cell(row=row, column=2, value=r["algorithm"])
            ws1.cell(row=row, column=3, value=r["seed"])
            ws1.cell(row=row, column=4, value=r["learning_rate"])
            ws1.cell(row=row, column=5, value=r["num_rounds"])
            ws1.cell(row=row, column=6, value=r["num_epochs"])
            ws1.cell(row=row, column=7, value=pct(r["max_val_acc"]))
            ws1.cell(row=row, column=8, value=r["best_val_round"])
            ws1.cell(row=row, column=9, value=pct(r["last_val_acc"]))
            ws1.cell(row=row, column=10, value=pct(r["val_gap"]))
            ws1.cell(row=row, column=11, value=pct(r["max_test_acc"]))
            ws1.cell(row=row, column=12, value=r["best_test_round"])
            ws1.cell(row=row, column=13, value=pct(r["last_test_acc"]))
            ws1.cell(row=row, column=14, value=pct(r["test_gap"]))
            ws1.cell(row=row, column=15, value=pct(r["test_at_best_val"]))
            ws1.cell(row=row, column=16, value=pct(r["mean_test_at_best_val"]))
            ws1.cell(row=row, column=17, value=pct(r["max_mean_test_acc"]))
            ws1.cell(row=row, column=18, value=r["param_sig"][:80])

            # Color-code val_gap
            gap_cell = ws1.cell(row=row, column=10)
            if r["val_gap"] is not None:
                if r["val_gap"] < 0.02:
                    gap_cell.fill = good_fill
                elif r["val_gap"] > 0.05:
                    gap_cell.fill = bad_fill
                else:
                    gap_cell.fill = neutral_fill

            row += 1

    auto_width(ws1)

    # ==========================================
    # Sheet 2: Per-Client Accuracy (at best val round)
    # ==========================================
    for task_name in ["PACS_c4", "office_caltech10_c4", "domainnet_c6"]:
        results = all_results.get(task_name, [])
        if not results:
            continue

        client_names = CLIENT_NAMES.get(task_name, [f"Client-{i}" for i in range(results[0]["num_clients"])])
        ws = wb.create_sheet(title=f"PerClient_{task_name[:10]}")

        headers = ["Algorithm", "Seed", "Max Val Acc%", "Best Round"] + \
                  [f"{c}%" for c in client_names] + \
                  ["Mean%", "Std%", "Min%", "Max%", "Param Signature"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, max_col=len(headers))

        row = 2
        results.sort(key=lambda x: (x["algorithm"], x.get("seed", 0) or 0))
        for r in results:
            ws.cell(row=row, column=1, value=r["algorithm"])
            ws.cell(row=row, column=2, value=r["seed"])
            ws.cell(row=row, column=3, value=pct(r["max_val_acc"]))
            ws.cell(row=row, column=4, value=r["best_val_round"])

            client_accs = r["per_client_acc_at_best"]
            for i, acc in enumerate(client_accs):
                ws.cell(row=row, column=5 + i, value=pct(acc))

            n_clients = len(client_names)
            if client_accs:
                arr = np.array(client_accs)
                ws.cell(row=row, column=5 + n_clients, value=pct(float(arr.mean())))
                ws.cell(row=row, column=6 + n_clients, value=pct(float(arr.std())))
                ws.cell(row=row, column=7 + n_clients, value=pct(float(arr.min())))
                ws.cell(row=row, column=8 + n_clients, value=pct(float(arr.max())))

            ws.cell(row=row, column=9 + n_clients, value=r["param_sig"][:60])
            row += 1

        auto_width(ws)

    # ==========================================
    # Sheet 3: Multi-Seed Summary (grouped by config)
    # ==========================================
    for task_name in ["PACS_c4", "office_caltech10_c4", "domainnet_c6"]:
        results = all_results.get(task_name, [])
        if not results:
            continue

        ws = wb.create_sheet(title=f"MultiSeed_{task_name[:10]}")

        headers = [
            "Algorithm", "Seeds", "N",
            "Max Val Mean±Std", "Max Test Mean±Std", "Test@BestVal Mean±Std",
            "Mean Test Mean±Std",
            "Last Val Mean±Std", "Last Test Mean±Std",
            "Val Gap Mean±Std", "Test Gap Mean±Std",
            "Best Round Mean", "Param Signature"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, max_col=len(headers))

        groups = group_by_config(results)
        row = 2
        # Sort by max_val_acc descending
        sorted_groups = sorted(groups.items(),
                               key=lambda x: -np.mean([r["max_val_acc"] or 0 for r in x[1]]))

        for sig, group in sorted_groups:
            stats = compute_multi_seed_stats(group)
            algo = group[0]["algorithm"]

            ws.cell(row=row, column=1, value=algo)
            ws.cell(row=row, column=2, value=str(stats["seeds"]))
            ws.cell(row=row, column=3, value=stats["n_seeds"])
            ws.cell(row=row, column=4, value=mean_std_str(stats["max_val_acc_mean"], stats["max_val_acc_std"]))
            ws.cell(row=row, column=5, value=mean_std_str(stats["max_test_acc_mean"], stats["max_test_acc_std"]))
            ws.cell(row=row, column=6, value=mean_std_str(stats["test_at_best_val_mean"], stats["test_at_best_val_std"]))
            ws.cell(row=row, column=7, value=mean_std_str(stats["mean_test_at_best_val_mean"], stats["mean_test_at_best_val_std"]))
            ws.cell(row=row, column=8, value=mean_std_str(stats["last_val_acc_mean"], stats["last_val_acc_std"]))
            ws.cell(row=row, column=9, value=mean_std_str(stats["last_test_acc_mean"], stats["last_test_acc_std"]))
            ws.cell(row=row, column=10, value=mean_std_str(stats["val_gap_mean"], stats["val_gap_std"]))
            ws.cell(row=row, column=11, value=mean_std_str(stats["test_gap_mean"], stats["test_gap_std"]))
            ws.cell(row=row, column=12, value=f"{stats['best_val_round_mean']:.0f}" if stats["best_val_round_mean"] else "")
            ws.cell(row=row, column=13, value=sig[:80])

            # Highlight multi-seed
            if stats["n_seeds"] >= 3:
                for c in range(1, 14):
                    ws.cell(row=row, column=c).fill = good_fill

            row += 1

        auto_width(ws)

    # ==========================================
    # Sheet 4: Experiment Catalog (EXP-001 to EXP-075)
    # ==========================================
    ws_cat = wb.create_sheet(title="EXP Catalog")

    headers = ["EXP ID", "Folder Name", "Category", "Status", "Description (from NOTE.md)"]
    for col, h in enumerate(headers, 1):
        ws_cat.cell(row=1, column=col, value=h)
    style_header(ws_cat, max_col=len(headers))

    row = 2
    for exp_num in range(1, 76):
        exp_id = f"EXP-{exp_num:03d}"
        if exp_num in exp_info:
            info = exp_info[exp_num]
            folder = info["folder_name"]
            # Determine category from folder path
            path = info["folder_path"]
            if "sanity" in path:
                cat = "sanity"
            elif "baselines" in path:
                cat = "baseline"
            elif "adaptive" in path:
                cat = "adaptive"
            elif "gradual" in path:
                cat = "gradual"
            else:
                cat = "ablation"

            # Extract first meaningful paragraph from NOTE
            content = info["note_full"]
            desc_lines = []
            in_desc = False
            for line in content.split("\n"):
                if line.startswith("## ") and "目标" in line or "说明" in line or "描述" in line or "Objective" in line or "Description" in line:
                    in_desc = True
                    continue
                if in_desc and line.startswith("## "):
                    break
                if in_desc and line.strip():
                    desc_lines.append(line.strip())

            if not desc_lines:
                # Fallback: first non-header lines
                for line in content.split("\n"):
                    if line.strip() and not line.startswith("#"):
                        desc_lines.append(line.strip())
                        if len(desc_lines) >= 3:
                            break

            desc = " ".join(desc_lines)[:300] if desc_lines else info["note_title"]

            ws_cat.cell(row=row, column=1, value=exp_id)
            ws_cat.cell(row=row, column=2, value=folder)
            ws_cat.cell(row=row, column=3, value=cat)
            ws_cat.cell(row=row, column=4, value="exists")
            ws_cat.cell(row=row, column=5, value=desc)
        else:
            ws_cat.cell(row=row, column=1, value=exp_id)
            ws_cat.cell(row=row, column=2, value="(not found)")
            ws_cat.cell(row=row, column=3, value="")
            ws_cat.cell(row=row, column=4, value="missing")
            ws_cat.cell(row=row, column=5, value="")
            for c in range(1, 6):
                ws_cat.cell(row=row, column=c).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        row += 1

    auto_width(ws_cat)
    ws_cat.column_dimensions['E'].width = 80

    # ==========================================
    # Sheet 5: Variant Descriptions
    # ==========================================
    ws_var = wb.create_sheet(title="Variant Descriptions")

    headers = ["Algorithm Variant", "Full Name", "Key Parameters", "Description",
               "PACS Runs", "Office Runs", "DomainNet Runs"]
    for col, h in enumerate(headers, 1):
        ws_var.cell(row=1, column=col, value=h)
    style_header(ws_var, max_col=len(headers))

    # Collect all unique variants
    variant_info = {}
    VARIANT_DESC = {
        "fedavg": "FedAvg baseline - standard federated averaging",
        "fedbn": "FedBN baseline - local BN layers, rest aggregated",
        "fedproto": "FedProto baseline - prototype-based FL",
        "fedprox": "FedProx baseline - proximal regularization (mu=0.1)",
        "moon": "MOON baseline - model contrastive FL (mu=0.1, tau=0.5)",
        "ditto": "Ditto baseline - personalized FL with local fine-tuning",
        "fdse": "FDSE (CVPR'25) - domain shift eraser with layer decomposition",
        "feddsa": "FedDSA - dual head decoupling + style repository + semantic alignment",
        "feddsa_adaptive": "FedDSA-Adaptive - with M0/M2/M3 adaptive modes (auto alpha, domain-aware protos)",
        "feddsa_augschedule": "FedDSA + augmentation scheduling (peak at 67% of rounds)",
        "feddsa_auto": "FedDSA-Auto - automatic loss weight tuning",
        "feddsa_cka": "FedDSA + CKA (Centered Kernel Alignment) independence loss",
        "feddsa_consensus": "FedDSA + consensus-aware QP aggregation (FDSE-style)",
        "feddsa_domain_aware": "FedDSA + domain-aware client grouping",
        "feddsa_fixbn": "FedDSA + fixed BN (cross-client BN stats alignment)",
        "feddsa_gated": "FedDSA + gated style dispatch (distance threshold=1.0)",
        "feddsa_gradual": "FedDSA-Gradual - sigmoid ramp-up + shallow augmentation",
        "feddsa_multilayer": "FedDSA + multi-layer style representation",
        "feddsa_noaug": "FedDSA without style augmentation (decoupling+alignment only)",
        "feddsa_noaug_late": "FedDSA - stop augmentation after round 150",
        "feddsa_pcgrad": "FedDSA + PCGrad gradient conflict resolution",
        "feddsa_regime_gated": "FedDSA + regime-aware gating v1",
        "feddsa_regime_gated_v2": "FedDSA + regime-aware gating v2 (improved)",
        "feddsa_softbeta": "FedDSA + soft Beta mixing for style augmentation",
        "feddsa_stylehead_bank": "FedDSA + bank of multiple style heads",
        "feddsa_triplet": "FedDSA + triplet loss for decoupling (margin=0.3)",
        "feddsa_vae": "FedDSA + VAE-based style head (KL divergence regularization)",
    }

    for task_name, results in all_results.items():
        for r in results:
            algo = r["algorithm"]
            if algo not in variant_info:
                variant_info[algo] = {
                    "algo": algo,
                    "desc": VARIANT_DESC.get(algo, ""),
                    "params": set(),
                    "pacs": 0, "office": 0, "domainnet": 0
                }
            variant_info[algo]["params"].add(r["param_sig"])
            if "PACS" in task_name:
                variant_info[algo]["pacs"] += 1
            elif "office" in task_name:
                variant_info[algo]["office"] += 1
            elif "domainnet" in task_name:
                variant_info[algo]["domainnet"] += 1

    row = 2
    for algo in sorted(variant_info.keys()):
        vi = variant_info[algo]
        ws_var.cell(row=row, column=1, value=algo)
        ws_var.cell(row=row, column=2, value=algo)
        # Show a sample parameter signature
        sample_params = sorted(vi["params"])
        ws_var.cell(row=row, column=3, value=sample_params[0][:80] if sample_params else "")
        ws_var.cell(row=row, column=4, value=vi["desc"])
        ws_var.cell(row=row, column=5, value=vi["pacs"])
        ws_var.cell(row=row, column=6, value=vi["office"])
        ws_var.cell(row=row, column=7, value=vi["domainnet"])
        row += 1

    auto_width(ws_var)
    ws_var.column_dimensions['D'].width = 60

    # ==========================================
    # Sheet 6: Best Results Ranking (sorted by weighted test acc)
    # ==========================================
    for task_name in ["PACS_c4", "office_caltech10_c4", "domainnet_c6"]:
        results = all_results.get(task_name, [])
        if not results:
            continue

        ws = wb.create_sheet(title=f"Ranking_{task_name[:10]}")

        headers = [
            "Rank", "Algorithm", "Seed", "Max Val Acc%",
            "Test@BestVal%", "Mean Test@BestVal%",
            "Best Round", "Last Val%", "Val Gap%", "Param Sig"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, max_col=len(headers))

        # Sort by test_at_best_val descending
        sorted_results = sorted(results,
                                key=lambda x: -(x["test_at_best_val"] or 0))

        for rank, r in enumerate(sorted_results, 1):
            ws.cell(row=rank+1, column=1, value=rank)
            ws.cell(row=rank+1, column=2, value=r["algorithm"])
            ws.cell(row=rank+1, column=3, value=r["seed"])
            ws.cell(row=rank+1, column=4, value=pct(r["max_val_acc"]))
            ws.cell(row=rank+1, column=5, value=pct(r["test_at_best_val"]))
            ws.cell(row=rank+1, column=6, value=pct(r["mean_test_at_best_val"]))
            ws.cell(row=rank+1, column=7, value=r["best_val_round"])
            ws.cell(row=rank+1, column=8, value=pct(r["last_val_acc"]))
            ws.cell(row=rank+1, column=9, value=pct(r["val_gap"]))
            ws.cell(row=rank+1, column=10, value=r["param_sig"][:60])

            # Top-3 green, bottom-3 red
            if rank <= 3:
                for c in range(1, 11):
                    ws.cell(row=rank+1, column=c).fill = good_fill
            elif rank > len(sorted_results) - 3:
                for c in range(1, 11):
                    ws.cell(row=rank+1, column=c).fill = bad_fill

        auto_width(ws)

    # ==========================================
    # Sheet 7: Detailed NOTE.md content per EXP
    # ==========================================
    ws_notes = wb.create_sheet(title="EXP Notes Detail")
    headers = ["EXP ID", "Folder", "NOTE.md Content (first 1500 chars)"]
    for col, h in enumerate(headers, 1):
        ws_notes.cell(row=1, column=col, value=h)
    style_header(ws_notes, max_col=len(headers))

    row = 2
    for exp_num in sorted(exp_info.keys()):
        info = exp_info[exp_num]
        ws_notes.cell(row=row, column=1, value=info["exp_id"])
        ws_notes.cell(row=row, column=2, value=info["folder_name"])
        # Truncate content for Excel
        content = info["note_full"][:1500].replace("\r", "")
        ws_notes.cell(row=row, column=3, value=content)
        ws_notes.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    ws_notes.column_dimensions['C'].width = 120
    ws_notes.column_dimensions['A'].width = 12
    ws_notes.column_dimensions['B'].width = 35

    # Save
    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


def main():
    print("=" * 60)
    print("FedDSA Experiment Comprehensive Analyzer")
    print("=" * 60)

    print("\n[1/3] Scanning experiment folders and NOTE.md files...")
    exp_info = scan_all_experiments()
    print(f"  Found {len(exp_info)} experiment folders with NOTE.md")

    print("\n[2/3] Parsing all JSON result files...")
    all_results = scan_all_json_results()
    total = sum(len(v) for v in all_results.values())
    print(f"  Total: {total} result files parsed")

    print("\n[3/3] Generating Excel report...")
    output_path = os.path.join(BASE_DIR, "experiment_results_comprehensive.xlsx")
    write_excel(all_results, exp_info, output_path)

    # Print quick summary
    print("\n" + "=" * 60)
    print("QUICK SUMMARY")
    print("=" * 60)
    for task_name, results in all_results.items():
        if not results:
            continue
        print(f"\n--- {task_name} ({len(results)} runs) ---")
        # Top 5 by test_at_best_val
        sorted_r = sorted(results, key=lambda x: -(x["test_at_best_val"] or 0))
        print("  Top 5 by Test@BestVal:")
        for i, r in enumerate(sorted_r[:5]):
            seed_str = str(r['seed']) if r['seed'] is not None else "?"
            print(f"    {i+1}. {r['algorithm']:>25s} s={seed_str:>4s}  "
                  f"val={pct(r['max_val_acc']):>6.2f}%  "
                  f"test={pct(r['test_at_best_val']):>6.2f}%  "
                  f"gap={pct(r['val_gap']):>5.2f}%  "
                  f"R={r['best_val_round']}")

        # Worst 3
        print("  Bottom 3:")
        for i, r in enumerate(sorted_r[-3:]):
            seed_str = str(r['seed']) if r['seed'] is not None else "?"
            v = pct(r['max_val_acc']) or 0
            t = pct(r['test_at_best_val']) or 0
            g = pct(r['val_gap']) or 0
            print(f"    {len(sorted_r)-2+i}. {r['algorithm']:>25s} s={seed_str:>4s}  "
                  f"val={v:>6.2f}%  test={t:>6.2f}%  gap={g:>5.2f}%")


if __name__ == "__main__":
    main()
