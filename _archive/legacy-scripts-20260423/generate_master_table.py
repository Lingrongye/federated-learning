# -*- coding: utf-8 -*-
"""
Generate a single master Excel table with ALL experiments (EXP-001~075),
color-coded by phase, with full metrics from JSON result files.
"""
import json, os, re, glob, sys, io
import numpy as np
from collections import defaultdict, OrderedDict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 1. EXP descriptions & phase classification
# ============================================================
EXP_CATALOG = {
    # --- Phase 0: Sanity ---
    1:  ("Sanity", "FedDSA Sanity Check", "PACS 10轮功能验证，发现4个bug"),
    2:  ("Sanity", "FedDSA Bug Fix Verify", "PACS 50轮，验证bug修复"),

    # --- Phase 1: Baselines ---
    3:  ("Baseline", "FedAvg (PFLlib)", "PFLlib框架FedAvg基线"),
    4:  ("Baseline", "FedBN (PFLlib)", "PFLlib框架FedBN基线"),
    5:  ("Baseline", "FedProto (PFLlib)", "PFLlib框架FedProto基线"),
    6:  ("Baseline", "FedDSA 500R", "FDSE框架 FedDSA AlexNet 500轮基线"),
    7:  ("Baseline", "FedAvg 500R", "FDSE框架 FedAvg AlexNet 500轮"),
    8:  ("Baseline", "FedBN 500R", "FDSE框架 FedBN AlexNet 500轮"),
    9:  ("Baseline", "FedProto 500R", "FDSE框架 FedProto AlexNet 500轮"),
    10: ("Baseline", "FDSE 500R", "FDSE框架 FDSE(DSE) AlexNet 500轮"),
    56: ("Baseline", "FedProx", "FedProx mu=0.1 基线补齐"),
    57: ("Baseline", "MOON", "MOON mu=0.1 tau=0.5 基线补齐"),

    # --- Phase 2: Weight Tuning ---
    11: ("Weight Tuning", "Low Weight", "降低辅助损失权重: orth=0.1, hsic=0.01, sem=0.5"),
    12: ("Weight Tuning", "No HSIC", "去掉HSIC约束: hsic=0"),
    13: ("Weight Tuning", "Long Warmup", "延长warmup到50轮"),
    14: ("Weight Tuning", "V4 Strong", "warmup=50 + 原始强权重(1.0/0.1/1.0)"),
    15: ("Weight Tuning", "FedDSA+ 3-Stage", "三阶段训练 + Sigmoid自适应权重"),
    16: ("Weight Tuning", "V4 seed=15", "V4配置seed=15方差验证"),
    17: ("Weight Tuning", "V4 No HSIC ★Best", "V4+hsic=0 → PACS 80.93% (当时最佳)"),
    18: ("Weight Tuning", "FedDSA+ Late Stage", "FedDSA+ stage1=80, stage2=150"),
    19: ("Weight Tuning", "V4 lr=0.05", "V4学习率降低到0.05"),
    20: ("Weight Tuning", "V4 tau=0.5", "InfoNCE温度提高到0.5"),
    21: ("Weight Tuning", "V4 orth=2.0", "正交约束强度翻倍"),
    22: ("Weight Tuning", "HSIC=0+lr=0.05+FastDecay", "稳定性优化组合"),
    23: ("Weight Tuning", "HSIC=0+EMA=0.9", "服务器端EMA聚合"),
    24: ("Weight Tuning", "Softer Aug Beta=0.5", "更温和的风格增强"),
    25: ("Weight Tuning", "No InfoNCE", "只保留CE+CE_aug+Orth"),

    # --- Phase 3: Loss/Gradient Variants ---
    28: ("Loss Variant", "Uncertainty Weight", "Kendall多任务不确定性权重"),
    29: ("Loss Variant", "PCGrad", "投影冲突梯度"),
    30: ("Loss Variant", "Triplet Loss", "Triplet替代InfoNCE"),
    31: ("Loss Variant", "CKA Loss", "CKA替代cos²正交"),
    32: ("Loss Variant", "PCGrad+orth=2", "三重组合: HSIC=0+orth=2+PCGrad"),
    33: ("Loss Variant", "PCGrad+warmup=80", "PCGrad+更长warmup"),

    # --- Phase 4: Multi-Seed Validation ---
    35: ("Multi-Seed", "EXP-017 s=15", "最佳配置seed=15验证"),
    36: ("Multi-Seed", "EXP-017 s=333", "最佳配置seed=333验证"),
    43: ("Multi-Seed", "FDSE 3-seed", "FDSE基线多seed验证"),
    44: ("Multi-Seed", "MultiLayer 3-seed", "多层风格多seed验证"),
    45: ("Multi-Seed", "VAE 3-seed", "VAE风格头多seed验证"),
    46: ("Multi-Seed", "FedDSA 5-seed", "5-seed对齐FDSE官方: [2,4388,15,333,967]"),
    49: ("Multi-Seed", "FDSE R200 s=2", "补齐FDSE R200第5个seed"),
    50: ("Multi-Seed", "FixBN 3-seed", "BN修复多seed验证"),

    # --- Phase 5: Architecture Variants ---
    40: ("Architecture", "MultiLayer Style", "AdaIN同时在mid层和final层注入"),
    41: ("Architecture", "VAE StyleHead", "style_head输出(mu,log_var)+KL loss"),
    42: ("Architecture", "Asymmetric Heads", "语义头3层+残差, 风格头1层+L2norm"),

    # --- Phase 6: Bug Fixes ---
    47: ("Bug Fix", "Aug Ramp-Down(A)+HardStop(D)", "loss_aug渐减 / R150后停止增强"),
    48: ("Bug Fix", "FixBN", "修复BN Running Stats聚合Bug"),

    # --- Phase 7: Cross-Dataset ---
    51: ("Cross-Dataset", "Office-Caltech10", "第二数据集验证 FedDSA vs FDSE"),
    52: ("Cross-Dataset", "LR Grid Search", "PACS+Office LR网格搜索(0.001~0.5)"),

    # --- Phase 8: Style Diagnostic ---
    58: ("Style Fix", "Detach-Style", "阻断orth梯度污染encoder"),
    59: ("Style Fix", "StyleHead→Bank", "修复style_head与style_bank脱节"),
    60: ("Style Fix", "Gated Dispatch", "距离门控风格分发"),
    61: ("Style Fix", "Style Diag Suite", "风格共享诊断实验套件(4种fix)"),

    # --- Phase 9: Aggregation & Regime ---
    64: ("Aggregation", "Consensus QP", "共识感知QP聚合(FDSE式)"),
    65: ("Aggregation", "DomainNet Test", "DomainNet验证regime-dependent假说"),
    66: ("Aggregation", "Consensus+KL", "共识聚合+KL一致性正则"),
    67: ("Aggregation", "RegimeGated v1", "Style Graph Dispatch + Regime-Gated Server SAM"),
    68: ("Aggregation", "RegimeGated v2", "Fix Signal Source + Farthest-K Dispatch"),

    # --- Phase 10: Hyperparam & Component ---
    69: ("Hyperparam", "Grid Search 6x", "6组超参组合网格搜索"),
    70: ("Hyperparam", "Component Ablation", "3模块逐一去除消融"),
    71: ("Hyperparam", "Domain-Aware Proto", "域感知多正例InfoNCE替代均值原型"),

    # --- Phase 11: Adaptive ---
    72: ("Adaptive", "Adaptive Baselines", "M0/M2/M3固定alpha+自适应增强+域感知原型"),
    73: ("Adaptive", "M1+M3 Full", "自适应增强+域感知原型完整组合"),
    74: ("Adaptive", "Tau Investigation", "训练崩溃根因: tau=0.1→崩溃, tau=0.2→恢复"),

    # --- Phase 12: Gradual & Advanced ---
    75: ("Gradual", "Gradual Training", "三重修复: Sigmoid ramp + 浅层增强 + 梯度诊断"),
}

# Ditto baseline (no dedicated EXP folder, use unused number)
EXP_CATALOG[34] = ("Baseline", "Ditto", "Ditto个性化基线 mu=0.0001(PACS)/0.01(Office)")

# Missing EXPs
for n in [26,27,37,38,39,53,54,55,62,63]:
    EXP_CATALOG[n] = ("(Skipped)", f"EXP-{n:03d} (未创建)", "编号跳过")

# 28b special
EXP_CATALOG[280] = ("Loss Variant", "Uncertainty+Clamp(28b)", "不确定性权重+log_sigma clamp修复")

# Phase colors (RGB hex, no #)
PHASE_COLORS = {
    "Sanity":       "E2EFDA",  # light green
    "Baseline":     "D6E4F0",  # light blue
    "Weight Tuning":"FFF2CC",  # light yellow
    "Loss Variant": "FCE4D6",  # light orange
    "Multi-Seed":   "E2D9F3",  # light purple
    "Architecture": "D5F5E3",  # mint green
    "Bug Fix":      "FADBD8",  # light pink
    "Cross-Dataset":"FDEBD0",  # peach
    "Style Fix":    "F5CBA7",  # darker peach
    "Aggregation":  "AED6F1",  # sky blue
    "Hyperparam":   "F9E79F",  # gold
    "Adaptive":     "D2B4DE",  # medium purple
    "Gradual":      "A3E4D7",  # teal
    "(Skipped)":    "D5D8DC",  # gray
}

# ============================================================
# 2. Parse ALL JSON result files
# ============================================================
TASK_DIRS = {
    "PACS": os.path.join(BASE, "FDSE_CVPR25", "task", "PACS_c4", "record"),
    "Office": os.path.join(BASE, "FDSE_CVPR25", "task", "office_caltech10_c4", "record"),
    "DomainNet": os.path.join(BASE, "FDSE_CVPR25", "task", "domainnet_c6", "record"),
}

CLIENT_NAMES = {
    "PACS": ["Art", "Cartoon", "Photo", "Sketch"],
    "Office": ["Amazon", "Caltech", "DSLR", "Webcam"],
    "DomainNet": ["Clip", "Info", "Paint", "Quick", "Real", "Sketch"],
}


def parse_json(filepath):
    with open(filepath, "r") as f:
        data = json.load(f)
    opt = data.get("option", {})
    basename = os.path.basename(filepath)

    r = {"filename": basename, "filepath": filepath}
    r["seed"] = opt.get("seed")
    r["lr"] = opt.get("learning_rate")
    r["rounds"] = opt.get("num_rounds")
    r["epochs"] = opt.get("num_epochs")
    r["algo_para"] = str(opt.get("algo_para", ""))

    # Algorithm name
    algo_prefixes = [
        "feddsa_adaptive", "feddsa_augschedule", "feddsa_auto",
        "feddsa_cka", "feddsa_consensus_kl", "feddsa_consensus",
        "feddsa_domain_aware", "feddsa_fixbn", "feddsa_gradual",
        "feddsa_multilayer", "feddsa_noaug_late", "feddsa_noaug",
        "feddsa_pcgrad", "feddsa_regime_gated_v2", "feddsa_regime_gated",
        "feddsa_softbeta", "feddsa_stylehead_bank", "feddsa_triplet",
        "feddsa_vae", "feddsa_plus", "feddsa_stable", "feddsa",
        "fdse", "fedavg", "fedbn", "fedproto", "fedprox", "moon", "ditto",
    ]
    name = basename.replace(".json", "")
    r["algorithm"] = name.split("_")[0]
    for p in algo_prefixes:
        if name.startswith(p + "_"):
            r["algorithm"] = p
            break

    # Extract key hyperparams from filename
    def extract_param(pattern, text, default=None):
        m = re.search(pattern, text)
        return m.group(1) if m else default

    r["lambda_orth"] = extract_param(r"lambda_orth([\d.]+)", basename)
    r["lambda_hsic"] = extract_param(r"lambda_hsic([\d.]+)", basename)
    r["lambda_sem"] = extract_param(r"lambda_sem([\d.]+)", basename)
    r["tau"] = extract_param(r"tau([\d.]+)", basename)
    r["warmup"] = extract_param(r"warmup_rounds?(\d+)", basename)
    r["proj_dim"] = extract_param(r"proj_dim(\d+)", basename)

    # Param signature for grouping
    m = re.search(r'_S\d+_LD', basename)
    r["param_sig"] = basename[:m.start()] if m else basename[:80]

    # Metrics
    lva = data.get("local_val_accuracy", [])
    lta = data.get("local_test_accuracy", [])
    mlta = data.get("mean_local_test_accuracy", [])

    r["n_rounds_actual"] = len(lva)

    if lva:
        arr = np.array(lva)
        r["max_val"] = float(arr.max())
        r["best_round"] = int(arr.argmax())
        r["last_val"] = float(arr[-1])
        r["val_gap"] = float(arr.max() - arr[-1])
    else:
        r["max_val"] = r["best_round"] = r["last_val"] = r["val_gap"] = None

    if lta:
        arr = np.array(lta)
        r["max_test"] = float(arr.max())
        r["last_test"] = float(arr[-1])
        r["test_gap"] = float(arr.max() - arr[-1])
        r["test_at_best"] = float(lta[r["best_round"]]) if r["best_round"] is not None and r["best_round"] < len(lta) else None
    else:
        r["max_test"] = r["last_test"] = r["test_gap"] = r["test_at_best"] = None

    if mlta:
        r["mean_test_at_best"] = float(mlta[r["best_round"]]) if r["best_round"] is not None and r["best_round"] < len(mlta) else None
    else:
        r["mean_test_at_best"] = None

    # Per-client at best round
    test_dist = data.get("local_test_accuracy_dist", [])
    if test_dist and r["best_round"] is not None and r["best_round"] < len(test_dist):
        r["client_accs"] = [float(x) for x in test_dist[r["best_round"]]]
    else:
        r["client_accs"] = []

    return r


def load_all_results():
    results = {}
    for task, dir_path in TASK_DIRS.items():
        if not os.path.exists(dir_path):
            continue
        files = sorted(glob.glob(os.path.join(dir_path, "*.json")))
        results[task] = []
        for f in files:
            try:
                r = parse_json(f)
                r["task"] = task
                results[task].append(r)
            except Exception as e:
                print(f"ERROR: {f}: {e}")
        print(f"  {task}: {len(results[task])} files")
    return results


# ============================================================
# 3. Build mapping: JSON -> EXP number
# ============================================================
def map_json_to_exp(results_by_task):
    """Heuristic mapping from JSON filename/params to EXP number."""
    # Build a lookup: for each JSON result, assign an EXP number
    # based on algorithm variant + key hyperparameters

    def make_key(r):
        """Create a unique key from algorithm + key params."""
        parts = [r["algorithm"]]
        for p in ["lambda_orth", "lambda_hsic", "lambda_sem", "tau", "warmup", "proj_dim"]:
            if r.get(p):
                parts.append(f"{p}={r[p]}")
        parts.append(f"lr={r['lr']}")
        parts.append(f"R={r['rounds']}")
        return "|".join(str(x) for x in parts)

    # Manual mapping rules based on NOTE.md analysis
    # Format: (algorithm, conditions_dict) -> exp_num
    rules = []

    # Baselines
    rules.append(("fedavg", {"rounds": 500}, None, 7))
    rules.append(("fedavg", {"rounds": 200}, None, 7))
    rules.append(("fedbn", {"rounds": 500}, None, 8))
    rules.append(("fedbn", {"rounds": 200}, None, 8))
    rules.append(("fedproto", {}, None, 9))
    rules.append(("fdse", {}, "PACS", 10))
    rules.append(("fdse", {}, "Office", 51))
    rules.append(("fdse", {}, "DomainNet", 65))
    rules.append(("ditto", {}, None, 10))  # Ditto = baseline
    rules.append(("fedprox", {}, None, 56))
    rules.append(("moon", {}, None, 57))

    # The full mapping is complex - let's use a simplified approach
    # Group by param_sig and assign EXPs based on algorithm variant
    all_mapped = []

    for task, runs in results_by_task.items():
        for r in runs:
            exp = None
            algo = r["algorithm"]
            lo = r.get("lambda_orth", "")
            lh = r.get("lambda_hsic", "")
            ls = r.get("lambda_sem", "")
            tau = r.get("tau", "")
            wu = r.get("warmup", "")
            lr = r.get("lr")
            rounds = r.get("rounds")
            pd = r.get("proj_dim", "128")

            # Baselines
            if algo == "fedavg":
                exp = 7 if rounds == 500 else 7
            elif algo == "fedbn":
                exp = 8 if rounds == 500 else 8
            elif algo == "fedproto":
                exp = 9
            elif algo == "fdse":
                if task == "DomainNet": exp = 65
                elif task == "Office": exp = 51
                else:
                    exp = 10 if rounds == 500 else 43
            elif algo == "ditto":
                exp = 34  # Ditto baseline
            elif algo == "fedprox":
                exp = 56
            elif algo == "moon":
                exp = 57

            # FedDSA variants
            elif algo == "feddsa":
                # Distinguish by params
                if lh == "0.1" and wu == "10":
                    exp = 6  # original FedDSA 500R
                elif lo == "0.1" and lh == "0.01" and ls == "0.5":
                    exp = 11  # low weight
                elif lo == "0.5" and lh == "0.0" and ls == "0.5" and wu == "10":
                    exp = 12  # no HSIC v2
                elif lo == "0.5" and lh == "0.05" and ls == "0.5" and wu == "50":
                    exp = 13  # long warmup
                elif lo == "1.0" and lh == "0.1" and ls == "1.0" and wu == "50" and rounds == 200:
                    exp = 14  # V4 strong
                elif lo == "1.0" and lh == "0.0" and ls == "1.0" and tau == "0.1" and wu == "50":
                    if lr == 0.05 and pd == "128":
                        exp = 52  # LR grid search lr=0.05
                    elif lr == 0.2:
                        exp = 52  # LR grid search lr=0.2
                    elif wu == "9999":
                        exp = 70  # component ablation (no aux)
                    elif pd == "256":
                        exp = 74  # 256d baseline
                    else:
                        exp = 17  # V4 no HSIC = best config
                elif lo == "1.0" and lh == "0.0" and ls == "0.0" and wu == "50":
                    exp = 70  # ablation: no semantic
                elif lo == "1.0" and lh == "0.0" and ls == "0.0" and wu == "9999":
                    exp = 70  # ablation: no aux at all
                elif lo == "1.0" and lh == "0.0" and ls == "1.0" and wu == "9999":
                    exp = 70  # ablation: no aug (warmup=9999)
                elif lo == "1.0" and lh == "0.0" and ls == "0.5":
                    exp = 69  # hyperparam grid (sem=0.5)
                elif lo == "1.0" and lh == "0.0" and ls == "2.0":
                    exp = 69  # hyperparam grid (sem=2.0)
                elif lo == "2.0" and lh == "0.0":
                    exp = 69  # hyperparam grid (orth=2.0, hsic=0)
                elif lo == "2.0" and lh == "0.1":
                    exp = 21  # strong orth
                elif lo == "0.5" and lh == "0.0" and ls == "1.0":
                    exp = 69  # hyperparam grid (orth=0.5)
                elif lo == "1.0" and lh == "0.0" and ls == "1.0" and tau == "0.05":
                    exp = 69  # tau=0.05
                elif lo == "1.0" and lh == "0.0" and ls == "1.0" and tau == "0.2" and pd == "128":
                    exp = 69  # tau=0.2
                elif lo == "1.0" and lh == "0.0" and ls == "1.0" and tau == "0.2" and pd == "256":
                    exp = 74  # 256d tau=0.2
                elif lo == "1.0" and lh == "0.1" and tau == "0.5":
                    exp = 20  # tau=0.5
                elif lr == 0.05 and lh == "0.1":
                    exp = 19  # lr=0.05 with HSIC
                elif lr == 0.05 and lh == "0.0":
                    exp = 22  # lr=0.05 fast decay
                else:
                    exp = 17  # default FedDSA

            elif algo == "feddsa_fixbn":
                exp = 48
            elif algo == "feddsa_pcgrad":
                exp = 29
            elif algo == "feddsa_triplet":
                exp = 30
            elif algo == "feddsa_cka":
                exp = 31
            elif algo == "feddsa_auto":
                exp = 28
            elif algo == "feddsa_multilayer":
                exp = 40
            elif algo == "feddsa_vae":
                exp = 41
            elif algo == "feddsa_noaug":
                exp = 60  # or 25 depending on context
            elif algo == "feddsa_noaug_late":
                exp = 47
            elif algo == "feddsa_stylehead_bank":
                exp = 59
            elif algo == "feddsa_augschedule":
                exp = 47
            elif algo == "feddsa_softbeta":
                exp = 61
            elif algo == "feddsa_gated":
                exp = 60
            elif algo == "feddsa_consensus":
                exp = 64
            elif algo == "feddsa_consensus_kl":
                exp = 66
            elif algo == "feddsa_domain_aware":
                exp = 71
            elif algo == "feddsa_regime_gated":
                exp = 67
            elif algo == "feddsa_regime_gated_v2":
                exp = 68
            elif algo == "feddsa_adaptive":
                # Check mode
                if "_md0_" in r["filename"]:
                    exp = 72  # M0 baseline
                elif "_md2_" in r["filename"]:
                    exp = 72  # M2
                elif "_md3_" in r["filename"]:
                    exp = 73  # M3
                else:
                    exp = 72

            r["exp_num"] = exp
            all_mapped.append(r)

    return all_mapped


# ============================================================
# 4. Multi-seed aggregation
# ============================================================
def aggregate_seeds(runs):
    """Group runs by (task, param_sig) and compute mean±std."""
    groups = defaultdict(list)
    for r in runs:
        key = (r["task"], r["param_sig"])
        groups[key].append(r)

    agg_rows = []
    for (task, sig), group in groups.items():
        row = dict(group[0])  # copy first run as template
        row["n_seeds"] = len(group)
        row["seeds_list"] = sorted([r["seed"] for r in group if r["seed"] is not None])

        for metric in ["max_val", "max_test", "test_at_best", "mean_test_at_best",
                        "last_val", "last_test", "val_gap", "test_gap", "best_round"]:
            vals = [r[metric] for r in group if r[metric] is not None]
            if vals:
                row[f"{metric}_mean"] = float(np.mean(vals))
                row[f"{metric}_std"] = float(np.std(vals))
            else:
                row[f"{metric}_mean"] = None
                row[f"{metric}_std"] = None

        # Per-client mean
        all_clients = [r["client_accs"] for r in group if r["client_accs"]]
        if all_clients and all(len(c) == len(all_clients[0]) for c in all_clients):
            arr = np.array(all_clients)
            row["client_means"] = arr.mean(axis=0).tolist()
            row["client_stds"] = arr.std(axis=0).tolist()
        else:
            row["client_means"] = group[0].get("client_accs", [])
            row["client_stds"] = []

        row["source_files"] = [r.get("filename", "") for r in group]
        agg_rows.append(row)

    return agg_rows


# ============================================================
# 5. Excel generation
# ============================================================
def pct(v):
    return round(v * 100, 2) if v is not None else ""

def ms(mean, std):
    if mean is None: return ""
    if std is not None and std > 0.001:
        return f"{mean*100:.2f}+-{std*100:.2f}"
    return f"{mean*100:.2f}"


def write_master_excel(all_runs, agg_rows, output_path):
    wb = openpyxl.Workbook()

    thin = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin')
    )
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    section_font = Font(bold=True, size=11, color="FFFFFF")

    def auto_width(ws, min_w=6, max_w=35):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            length = min_w
            for cell in col:
                if cell.value:
                    length = max(length, min(len(str(cell.value)), max_w))
            ws.column_dimensions[letter].width = length + 2

    # ============================
    # Per-dataset master sheets
    # ============================
    TASK_META = OrderedDict([
        ("PACS",      {"tab_color": "2F5496", "clients": ["Art", "Cartoon", "Photo", "Sketch"]}),
        ("Office",    {"tab_color": "548235", "clients": ["Amazon", "Caltech", "DSLR", "Webcam"]}),
        ("DomainNet", {"tab_color": "BF8F00", "clients": ["Clipart", "Infograph", "Painting", "Quickdraw", "Real", "Sketch"]}),
    ])

    first_sheet = True
    for task_key, meta in TASK_META.items():
        task_runs = [r for r in all_runs if r["task"] == task_key]
        if not task_runs:
            continue

        clients = meta["clients"]
        n_clients = len(clients)

        if first_sheet:
            ws = wb.active
            ws.title = task_key
            first_sheet = False
        else:
            ws = wb.create_sheet(task_key)
        ws.sheet_properties.tabColor = meta["tab_color"]

        headers = (
            ["EXP#", "Phase", "Variant Name", "Description",
             "Algorithm", "Seed",
             "Max Val%", "Best Round", "Last Val%", "Val Gap%",
             "Test@Best%", "Mean Test@Best%", "Max Test%", "Last Test%", "Test Gap%"]
            + [f"{c}%" for c in clients]
            + ["LR", "Rounds", "JSON Source"]
        )
        n_cols = len(headers)
        # Col index where client columns start (1-based)
        client_col_start = 16
        gap_col = 10  # Val Gap%

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin
        ws.freeze_panes = "A2"

        # Sort by EXP number then seed
        task_runs.sort(key=lambda r: (r.get("exp_num") or 999, r.get("seed") or 0))

        row = 2
        current_phase = None

        for r in task_runs:
            exp_num = r.get("exp_num")
            if exp_num and exp_num in EXP_CATALOG:
                phase, vname, desc = EXP_CATALOG[exp_num]
            else:
                phase, vname, desc = "Other", r["algorithm"], ""

            # Phase separator row
            if phase != current_phase:
                current_phase = phase
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = PatternFill("solid", fgColor="404040")
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.border = thin
                ws.cell(row=row, column=1, value=f"▶ {phase}")
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_cols)
                ws.cell(row=row, column=2, value=phase)
                row += 1

            # Data row
            phase_color = PHASE_COLORS.get(phase, "FFFFFF")
            fill = PatternFill("solid", fgColor=phase_color)

            vals = [
                f"EXP-{exp_num:03d}" if exp_num else "",
                phase, vname, desc[:50],
                r["algorithm"], r.get("seed"),
                pct(r["max_val"]), r["best_round"], pct(r["last_val"]), pct(r["val_gap"]),
                pct(r["test_at_best"]), pct(r["mean_test_at_best"]),
                pct(r["max_test"]), pct(r["last_test"]), pct(r["test_gap"]),
            ]
            ca = r.get("client_accs", [])
            for i in range(n_clients):
                vals.append(pct(ca[i]) if i < len(ca) else "")
            vals.extend([r.get("lr"), r.get("rounds"), r.get("filename", "")])

            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(horizontal='center' if c <= client_col_start + n_clients else 'left',
                                           vertical='center')

            # Highlight val gap
            gc = ws.cell(row=row, column=gap_col)
            if r["val_gap"] is not None:
                if r["val_gap"] > 0.05:
                    gc.font = Font(color="FF0000", bold=True)
                elif r["val_gap"] < 0.02:
                    gc.font = Font(color="006100", bold=True)

            row += 1

        auto_width(ws)
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions[get_column_letter(n_cols)].width = 80  # JSON Source

    # ============================
    # Per-dataset multi-seed summary sheets
    # ============================
    for task_key, meta in TASK_META.items():
        task_agg = [r for r in agg_rows if r["task"] == task_key]
        if not task_agg:
            continue

        clients = meta["clients"]
        n_clients = len(clients)

        ws2 = wb.create_sheet(f"Summary_{task_key}")
        ws2.sheet_properties.tabColor = meta["tab_color"]

        headers2 = (
            ["EXP#", "Phase", "Variant", "Algorithm", "Seeds", "N",
             "MaxVal Mean+-Std", "Test@Best Mean+-Std", "MeanTest Mean+-Std",
             "LastVal Mean+-Std", "ValGap Mean+-Std", "BestRound Mean"]
            + [f"{c} Mean+-Std" for c in clients]
            + ["JSON Sources"]
        )

        for c, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor=meta["tab_color"])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin
        ws2.freeze_panes = "A2"

        task_agg.sort(key=lambda r: -(r.get("test_at_best_mean") or 0))

        row2 = 2
        for r in task_agg:
            exp_num = r.get("exp_num")
            if exp_num and exp_num in EXP_CATALOG:
                phase, vname, _ = EXP_CATALOG[exp_num]
            else:
                phase, vname = "Other", r["algorithm"]

            fill = PatternFill("solid", fgColor=PHASE_COLORS.get(phase, "FFFFFF"))

            vals2 = [
                f"EXP-{exp_num:03d}" if exp_num else "",
                phase, vname, r["algorithm"],
                str(r.get("seeds_list", [])), r.get("n_seeds", 1),
                ms(r.get("max_val_mean"), r.get("max_val_std")),
                ms(r.get("test_at_best_mean"), r.get("test_at_best_std")),
                ms(r.get("mean_test_at_best_mean"), r.get("mean_test_at_best_std")),
                ms(r.get("last_val_mean"), r.get("last_val_std")),
                ms(r.get("val_gap_mean"), r.get("val_gap_std")),
                f"{r.get('best_round_mean', 0):.0f}" if r.get("best_round_mean") else "",
            ]
            cm = r.get("client_means", [])
            cs = r.get("client_stds", [])
            for i in range(n_clients):
                if i < len(cm):
                    vals2.append(ms(cm[i], cs[i] if i < len(cs) and cs else 0))
                else:
                    vals2.append("")
            srcs = r.get("source_files", [])
            vals2.append(" ; ".join(srcs))

            for c, v in enumerate(vals2, 1):
                cell = ws2.cell(row=row2, column=c, value=v)
                cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(horizontal='center')

            if r.get("n_seeds", 1) >= 3:
                for c in range(1, 7):
                    ws2.cell(row=row2, column=c).font = Font(bold=True)
            row2 += 1

        auto_width(ws2)
        ws2.column_dimensions[get_column_letter(len(headers2))].width = 100  # JSON Sources

    # ============================
    # SHEET 3: EXP Catalog (001-075)
    # ============================
    ws3 = wb.create_sheet("EXP Catalog 001-075")
    ws3.sheet_properties.tabColor = "BF8F00"

    headers3 = ["EXP#", "Phase", "Variant Name", "Description",
                 "PACS Runs", "Office Runs", "DN Runs", "Status"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="BF8F00")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin

    # Count runs per EXP per task
    exp_counts = defaultdict(lambda: {"PACS": 0, "Office": 0, "DomainNet": 0})
    for r in all_runs:
        if r.get("exp_num"):
            exp_counts[r["exp_num"]][r["task"]] += 1

    row3 = 2
    for n in range(1, 76):
        if n in EXP_CATALOG:
            phase, vname, desc = EXP_CATALOG[n]
        else:
            phase, vname, desc = "(Unknown)", f"EXP-{n:03d}", ""

        color = PHASE_COLORS.get(phase, "FFFFFF")
        fill = PatternFill("solid", fgColor=color)

        counts = exp_counts.get(n, {"PACS": 0, "Office": 0, "DomainNet": 0})
        total = sum(counts.values())
        status = "Has Data" if total > 0 else ("Skipped" if phase == "(Skipped)" else "No JSON")

        vals3 = [
            f"EXP-{n:03d}", phase, vname, desc,
            counts["PACS"], counts["Office"], counts["DomainNet"], status
        ]
        for c, v in enumerate(vals3, 1):
            cell = ws3.cell(row=row3, column=c, value=v)
            cell.fill = fill
            cell.border = thin

        row3 += 1

    auto_width(ws3)
    ws3.column_dimensions['D'].width = 50

    # Save
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"  Sheet 1 'Master Table': {ws.max_row} rows")
    print(f"  Sheet 2 'Multi-Seed Summary': {ws2.max_row} rows")
    print(f"  Sheet 3 'EXP Catalog': {ws3.max_row} rows")


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("Master Table Generator — FedDSA All Experiments")
    print("=" * 60)

    print("\n[1/4] Loading all JSON results...")
    results_by_task = load_all_results()
    total = sum(len(v) for v in results_by_task.values())
    print(f"  Total: {total} JSON files")

    print("\n[2/4] Mapping to EXP numbers...")
    all_runs = map_json_to_exp(results_by_task)
    mapped = sum(1 for r in all_runs if r.get("exp_num"))
    print(f"  Mapped: {mapped}/{len(all_runs)}")

    print("\n[3/4] Aggregating multi-seed runs...")
    agg_rows = aggregate_seeds(all_runs)
    print(f"  Unique configs: {len(agg_rows)}")

    print("\n[4/4] Writing Excel...")
    out = os.path.join(BASE, "experiment_master_table_v3.xlsx")
    write_master_excel(all_runs, agg_rows, out)

    # Quick stats
    print("\n" + "=" * 60)
    print("TOP 10 by Test@BestVal (PACS)")
    pacs = [r for r in all_runs if r["task"] == "PACS"]
    pacs.sort(key=lambda x: -(x.get("test_at_best") or 0))
    for i, r in enumerate(pacs[:10]):
        seed = str(r.get("seed", "?"))
        exp = str(r.get("exp_num", "?"))
        tv = pct(r["test_at_best"]) if r["test_at_best"] else 0
        vv = pct(r["max_val"]) if r["max_val"] else 0
        gv = pct(r["val_gap"]) if r["val_gap"] else 0
        print(f"  {i+1:>2}. EXP-{exp:>3s} {r['algorithm']:>25s} s={seed:>4s}  "
              f"test={tv:>6}%  val={vv:>6}%  gap={gv:>5}%  R={r['best_round']}")


if __name__ == "__main__":
    main()
